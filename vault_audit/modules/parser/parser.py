import openpyxl
from modules.models.models import Parameters, Transaction, ContainerData
from datetime import datetime, date
import re

FILTERED_LABELS = {
    "Bags",
    "Labels",
    "Non-std  : Pennies",
    "Non-std  : Dimes",
    "Non-std bags : Pennies",
    "Non-std bags : Nickels",
    "Non-std bags : Nickles",
    "Non-std bags : Dimes",
    "Non-std bags : Quarters",
    "Non-std bags : Dollars",
    "Boxes : Pennies",
    "Boxes : Nickels",
    "Boxes : Dimes",
    "Boxes : Quarters",
    "Boxes : Half dollars",
    "Boxes : Dollars",
    "Bags : Pennies",
    "Bags : Nickels",
    "Bags : Nickles",
    "Bags : Dimes",
    "Bags : Quarters",
    "Bags : Half dollars",
    "Bags : Dollars"
}


def parse_parameters(sheet) -> Parameters:
    carrier_location_full = str(sheet['B4'].value) if sheet['B4'].value else ""

    carrier_location = carrier_location_full
    if " : " in carrier_location_full:
        parts = carrier_location_full.split(" : ", 1)
        if len(parts) == 2:
            carrier_location = parts[1]

    created_at_raw = sheet['B1'].value
    created_at_formatted = str(created_at_raw) if created_at_raw else ""
    created_at_date_obj = None

    # Try to extract date object for import tracking
    try:
        # If B1 is already a datetime object (from openpyxl)
        if isinstance(created_at_raw, datetime):
            created_at_date_obj = created_at_raw.date()
            created_at_formatted = created_at_raw.strftime("%m/%d/%y %H:%M:%S")
        # If B1 is a string, parse it
        elif isinstance(created_at_raw, str):
            # Try format: "2025-10-11 09:00 AM CDT"
            match = re.match(r'(\d{4})-(\d{2})-(\d{2})\s+(\d{2}):(\d{2})\s+(AM|PM)\s+(\w+)', created_at_raw)
            if match:
                year, month, day, hour, minute, ampm, tz = match.groups()
                hour = int(hour)

                if ampm == 'PM' and hour != 12:
                    hour += 12
                elif ampm == 'AM' and hour == 12:
                    hour = 0

                dt = datetime(int(year), int(month), int(day), hour, int(minute))
                created_at_date_obj = dt.date()
                created_at_formatted = dt.strftime(f"%m/%d/%y %H:%M:%S {tz}")
            else:
                # Try simpler format: "2025-10-11 00:00:00" or "2025-10-11"
                simple_match = re.match(r'(\d{4})-(\d{2})-(\d{2})', created_at_raw)
                if simple_match:
                    year, month, day = simple_match.groups()
                    dt = datetime(int(year), int(month), int(day))
                    created_at_date_obj = dt.date()
                    created_at_formatted = dt.strftime("%m/%d/%y %H:%M:%S")
    except:
        pass

    # Fallback to today's date if parsing failed
    if not created_at_date_obj:
        created_at_date_obj = date.today()

    return Parameters(
        created_at=created_at_formatted,
        created_by=sheet['B2'].value,
        carrier=sheet['B3'].value,
        carrier_location=carrier_location,
        created_at_date=created_at_date_obj
    )


def parse_dynamic_sheet(sheet) -> tuple[list[Transaction], set[str]]:
    transactions = []
    valid_labels = set()

    current_transaction = None
    current_labels = []
    current_count = 0.0
    current_value = 0.0

    for row_idx in range(2, sheet.max_row + 1):
        origin = sheet.cell(row_idx, 1).value
        destination = sheet.cell(row_idx, 2).value
        trans_type = sheet.cell(row_idx, 3).value
        departure = sheet.cell(row_idx, 4).value
        arrival = sheet.cell(row_idx, 5).value
        label = sheet.cell(row_idx, 6).value
        count = sheet.cell(row_idx, 7).value
        value = sheet.cell(row_idx, 8).value

        origin_str = str(origin).strip() if origin else ""
        destination_str = str(destination).strip() if destination else ""
        trans_type_str = str(trans_type).strip() if trans_type else ""

        is_primary_row = origin_str and destination_str and trans_type_str
        is_date_separator = origin_str and not destination_str and not trans_type_str

        if is_date_separator:
            if current_transaction:
                transactions.append(current_transaction)
                current_transaction = None
            continue

        if is_primary_row:
            if current_transaction:
                transactions.append(current_transaction)

            current_labels = []
            current_count = 0.0
            current_value = 0.0

            current_transaction = Transaction(
                origin=origin_str,
                destination=destination_str,
                type=trans_type_str,
                departure_date=str(departure).strip() if departure else "",
                arrival_date=str(arrival).strip() if arrival else "",
                labels=[],
                total_count=0.0,
                total_value=0.0
            )

        if label:
            label_str = str(label).strip()
            if label_str and label_str not in FILTERED_LABELS:
                valid_labels.add(label_str)
                if current_transaction:
                    current_labels.append(label_str)

        if count:
            current_count += float(count)
        if value:
            current_value += float(value)

        if current_transaction:
            current_transaction.labels = current_labels[:]
            current_transaction.total_count = current_count
            current_transaction.total_value = current_value

    if current_transaction:
        transactions.append(current_transaction)

    return transactions, valid_labels


def parse_container_file(file_path: str) -> ContainerData:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    params_sheet = wb['Parameters']
    parameters = parse_parameters(params_sheet)

    location_sheet_name = wb.sheetnames[1]
    location_sheet = wb[location_sheet_name]

    transactions, valid_labels = parse_dynamic_sheet(location_sheet)

    return ContainerData(
        parameters=parameters,
        location_name=location_sheet_name,
        valid_labels=valid_labels,
        transactions=transactions
    )