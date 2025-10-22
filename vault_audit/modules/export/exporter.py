import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from datetime import datetime
import os
from modules.models.models import AuditResult


def export_audit_results(audit_result: AuditResult, container_info: dict, output_folder: str = "exports", bag_durations: dict = None, labels_over_3_days: list = None) -> str:
    os.makedirs(output_folder, exist_ok=True)

    wb = openpyxl.Workbook()

    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    summary_sheet['A1'] = "Vault Audit Report"
    summary_sheet['A1'].font = Font(size=16, bold=True)
    summary_sheet['A1'].alignment = Alignment(horizontal='left')

    summary_sheet['A3'] = "Report Generated:"
    summary_sheet['B3'] = datetime.now().strftime("%m/%d/%y %H:%M:%S CST")

    summary_sheet['A4'] = "Container Holdover Date:"
    summary_sheet['B4'] = container_info.get('created_at', 'N/A')

    summary_sheet['A5'] = "Location:"
    summary_sheet['B5'] = container_info.get('location', 'N/A')

    # Audit Summary section
    audit_summary_row = 7

    summary_sheet[f'A{audit_summary_row}'] = "Audit Summary"
    summary_sheet[f'A{audit_summary_row}'].font = Font(size=14, bold=True)

    location = container_info.get('location', 'N/A')

    # Calculate dynamic row positions
    containers_row = audit_summary_row + 2
    scanned_row = containers_row + 1
    matched_row = scanned_row + 1
    unmatched_row = matched_row + 1
    not_scanned_row = unmatched_row + 1

    summary_sheet[f'A{containers_row}'] = "Total Containers in Onsite:"
    summary_sheet[f'B{containers_row}'] = len(audit_result.expected_not_scanned) + len(audit_result.matched_labels)
    summary_sheet[f'B{containers_row}'].font = Font(bold=True)
    summary_sheet[f'A{containers_row}'].comment = Comment(f"Total Containers in {location} onsite", "System")

    summary_sheet[f'A{scanned_row}'] = "Total Scanned:"
    summary_sheet[f'B{scanned_row}'] = audit_result.total_scanned
    summary_sheet[f'B{scanned_row}'].font = Font(bold=True)

    # Calculate count of bags >=3 days (use import-based tracking if available)
    bags_3_plus_days = 0
    if labels_over_3_days:
        bags_3_plus_days = len(labels_over_3_days)
    elif bag_durations:
        bags_3_plus_days = sum(1 for label in audit_result.matched_labels
                               if bag_durations.get(label) and bag_durations[label]['days_in_vault'] >= 3)

    summary_sheet[f'A{matched_row}'] = "Labels >=3 Days in Vault:"
    summary_sheet[f'B{matched_row}'] = f"🔥 {bags_3_plus_days}"
    summary_sheet[f'B{matched_row}'].font = Font(color="FF0000", bold=True)
    summary_sheet[f'A{matched_row}'].comment = Comment(f"Bags that have been in container-holdover for 3 or more days (import-based tracking)", "System")

    summary_sheet[f'A{unmatched_row}'] = "Unmatched Labels:"
    summary_sheet[f'B{unmatched_row}'] = len(audit_result.unmatched_labels)
    summary_sheet[f'B{unmatched_row}'].font = Font(color="FF0000", bold=True)
    summary_sheet[f'A{unmatched_row}'].comment = Comment(f"Physical Bag Found but not in {location} onsite", "System")

    summary_sheet[f'A{not_scanned_row}'] = "Not Scanned:"
    summary_sheet[f'B{not_scanned_row}'] = len(audit_result.expected_not_scanned)
    summary_sheet[f'B{not_scanned_row}'].font = Font(color="FFA500", bold=True)
    summary_sheet[f'A{not_scanned_row}'].comment = Comment(f"Bags in {location} onsite but physically in the {location} vault", "System")

    for col in ['A', 'B']:
        summary_sheet.column_dimensions[col].width = 25

    results_sheet = wb.create_sheet("Results")

    # Simple vertical layout - Section 1: BAGS IN VAULT 3+ DAYS
    row = 1
    results_sheet[f'A{row}'] = "BAGS IN VAULT 3+ DAYS"
    results_sheet[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    results_sheet[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    results_sheet[f'A{row}'].alignment = Alignment(horizontal='left')

    row += 1
    results_sheet[f'A{row}'] = "Label ID"
    results_sheet[f'A{row}'].font = Font(bold=True)
    results_sheet[f'B{row}'] = "First Seen"
    results_sheet[f'B{row}'].font = Font(bold=True)
    results_sheet[f'C{row}'] = "Days"
    results_sheet[f'C{row}'].font = Font(bold=True)

    row += 1
    if labels_over_3_days:
        for label_info in labels_over_3_days:
            label = label_info['label_id']
            days = label_info['days_in_vault']
            first_import = label_info.get('first_import_date', 'Unknown')

            results_sheet[f'A{row}'] = label
            results_sheet[f'A{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            results_sheet[f'B{row}'] = first_import
            results_sheet[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            results_sheet[f'B{row}'].alignment = Alignment(horizontal='center')

            results_sheet[f'C{row}'] = days
            results_sheet[f'C{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            results_sheet[f'C{row}'].alignment = Alignment(horizontal='center')
            results_sheet[f'C{row}'].font = Font(bold=True)

            row += 1
    elif bag_durations:
        for label in sorted(audit_result.matched_labels):
            duration_info = bag_durations.get(label)
            if duration_info and duration_info['days_in_vault'] >= 3:
                first_scan = duration_info.get('first_scan')
                first_scan_date = first_scan.strftime('%Y-%m-%d') if first_scan else 'Unknown'
                days = duration_info['days_in_vault']

                results_sheet[f'A{row}'] = label
                results_sheet[f'A{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

                results_sheet[f'B{row}'] = first_scan_date
                results_sheet[f'B{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                results_sheet[f'B{row}'].alignment = Alignment(horizontal='center')

                results_sheet[f'C{row}'] = days
                results_sheet[f'C{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                results_sheet[f'C{row}'].alignment = Alignment(horizontal='center')
                results_sheet[f'C{row}'].font = Font(bold=True)

                row += 1

    if row == 3:  # No bags found
        results_sheet[f'A{row}'] = "None"
        results_sheet[f'A{row}'].font = Font(italic=True, color="666666")
        row += 1

    # Blank row separator
    row += 2

    # Section 2: UNMATCHED LABELS
    results_sheet[f'A{row}'] = "UNMATCHED LABELS"
    results_sheet[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    results_sheet[f'A{row}'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    results_sheet[f'A{row}'].alignment = Alignment(horizontal='left')

    row += 1
    results_sheet[f'A{row}'] = "Label ID"
    results_sheet[f'A{row}'].font = Font(bold=True)

    row += 1
    if audit_result.unmatched_labels:
        for label in sorted(audit_result.unmatched_labels):
            results_sheet[f'A{row}'] = label
            row += 1
    else:
        results_sheet[f'A{row}'] = "None"
        results_sheet[f'A{row}'].font = Font(italic=True, color="666666")
        row += 1

    # Blank row separator
    row += 2

    # Section 3: NOT SCANNED
    results_sheet[f'A{row}'] = "NOT SCANNED"
    results_sheet[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    results_sheet[f'A{row}'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    results_sheet[f'A{row}'].alignment = Alignment(horizontal='left')

    row += 1
    results_sheet[f'A{row}'] = "Label ID"
    results_sheet[f'A{row}'].font = Font(bold=True)

    row += 1
    if audit_result.expected_not_scanned:
        for label in sorted(audit_result.expected_not_scanned):
            results_sheet[f'A{row}'] = label
            row += 1
    else:
        results_sheet[f'A{row}'] = "None"
        results_sheet[f'A{row}'].font = Font(italic=True, color="666666")

    # Set column widths
    results_sheet.column_dimensions['A'].width = 40
    results_sheet.column_dimensions['B'].width = 20
    results_sheet.column_dimensions['C'].width = 15

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"vault_audit_report_{timestamp}.xlsx"
    filepath = os.path.join(output_folder, filename)

    wb.save(filepath)

    return filepath